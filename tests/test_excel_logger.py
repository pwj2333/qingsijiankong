from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from alarm_auto_handler.excel_logger import ExcelAlarmLogger


class ExcelAlarmLoggerTests(unittest.TestCase):
    def test_recreates_invalid_workbook_and_keeps_writing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            log_path = Path(tmp_dir) / "alarm_handle_log.xlsx"
            log_path.write_text("not a real xlsx", encoding="utf-8")

            logger = ExcelAlarmLogger(log_path)
            logger.append_row(
                event_type="告警处理",
                result="成功",
                message="测试写入",
                alarm_id="123",
            )

            workbook = load_workbook(log_path)
            sheet = workbook.active

            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(row=2, column=2).value, "告警处理")
            self.assertEqual(sheet.cell(row=2, column=4).value, "123")

            broken_files = list(log_path.parent.glob("alarm_handle_log.xlsx.broken.*"))
            self.assertEqual(len(broken_files), 1)


if __name__ == "__main__":
    unittest.main()

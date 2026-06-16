from __future__ import annotations

import logging
import shutil
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ExcelAlarmLogger:
    HEADERS = [
        "记录时间",
        "事件类型",
        "结果",
        "告警ID",
        "船名",
        "异常行为",
        "风险等级",
        "报警时间",
        "处理方式",
        "是否有效",
        "说明",
    ]

    _RECOVERABLE_EXCEPTIONS = (BadZipFile, InvalidFileException)

    def __init__(self, file_path: Path) -> None:
        self.file_path = file_path
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self._logger = logging.getLogger(__name__)
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        # If the workbook is damaged or not really an xlsx file, back it up and recreate it.
        if self.file_path.exists():
            try:
                load_workbook(self.file_path)
                return
            except self._RECOVERABLE_EXCEPTIONS:
                bak_path = self.file_path.with_suffix(
                    self.file_path.suffix + ".broken." + datetime.now().strftime("%Y%m%d%H%M%S")
                )
                try:
                    shutil.move(str(self.file_path), str(bak_path))
                except Exception:
                    try:
                        self.file_path.unlink()
                    except Exception:
                        pass

        self._create_workbook()

    def _create_workbook(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "处理日志"
        sheet.append(self.HEADERS)
        for index, width in enumerate((20, 14, 14, 24, 18, 26, 12, 20, 16, 12, 50), start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(self.file_path)

    def _load_workbook(self):
        try:
            return load_workbook(self.file_path)
        except self._RECOVERABLE_EXCEPTIONS:
            self._ensure_workbook()
            return load_workbook(self.file_path)

    def append_row(
        self,
        *,
        event_type: str,
        result: str,
        message: str,
        alarm_id: str = "",
        ship_name: str = "",
        behavior: str = "",
        risk_level: str = "",
        alarm_time: str = "",
        handle_method: str = "",
        is_valid: str = "",
    ) -> None:
        try:
            workbook = self._load_workbook()
            sheet = workbook.active
            sheet.append(
                [
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    event_type,
                    result,
                    alarm_id,
                    ship_name,
                    behavior,
                    risk_level,
                    alarm_time,
                    handle_method,
                    is_valid,
                    message,
                ]
            )
            workbook.save(self.file_path)
        except Exception:
            self._logger.exception("写入 Excel 日志失败: %s", self.file_path)
